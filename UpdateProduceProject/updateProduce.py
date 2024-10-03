#! python3
# updateProduce.py - Корректирует затраты в таблице продаж продуктов.

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')

sheet = wb['Sheet']

# Типы продуктов и из обновленные цены
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

# Пройти по строкам и обновить цены
for rowNum in range(2, sheet.max_row): # пропустить первую строку
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')

